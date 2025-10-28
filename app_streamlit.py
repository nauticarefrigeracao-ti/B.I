import sqlite3
import os
import requests
from pathlib import Path
import pandas as pd
import streamlit as st
from datetime import datetime, timezone
import importlib.util
import html
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl.styles import Font, Alignment, numbers
import base64
import json

DT_CSS = "https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css"
DT_JS = "https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"
JQ = "https://code.jquery.com/jquery-3.5.1.js"

def render_interactive_table(df, table_id='tbl'):
    """Return HTML snippet for an interactive DataTable (client-side)."""
    if df is None or df.empty:
        return '<div>(vazio)</div>'

    # Work on a copy and normalize column names for display
    df2 = df.copy()
    df2.columns = [c.replace('_', ' ').strip() for c in df2.columns]

    # format datetimes
    for c in df2.select_dtypes(include=['datetime', 'datetimetz']).columns:
        df2[c] = df2[c].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')

    # if there's an index column '#' and an order id column, convert the index to HTML
    cols_lower = [col.lower() for col in df2.columns]
    if '#' in df2.columns and 'order id' in cols_lower:
        html_idx = []
        # find the actual column name for order id after normalization
        oid_col = next(col for col in df2.columns if col.lower() == 'order id')
        for i, oid in enumerate(df2[oid_col].fillna('').astype(str), start=1):
            # Keep the index number for identification/organization. We used
            # to render a gold copy button here but it's redundant with the
            # copy action included in the Order ID column. Remove it to save
            # horizontal space.
            idx_html = f'<span class="row-index" title="Linha {i}">{i}</span>'
            html_idx.append(idx_html)
        df2['#'] = html_idx

    # Make the Order ID itself a clickable link to Mercado Livre detail page so the full ID is visible and can be opened
    if 'order id' in cols_lower:
        order_col_name = next(col for col in df2.columns if col.lower() == 'order id')
        df2[order_col_name] = df2[order_col_name].fillna('').astype(str).apply(
                # Build a compact action group: main link to MercadoLivre detail (opens in new tab),
                # a copy button, a button that navigates the top window to prefill the review form,
                # and a button that navigates the top window to open the detail view inside the app.
                lambda oid: (
                    (f'<a href="https://www.mercadolivre.com.br/vendas/{html.escape(oid)}/detalhe" target="_blank" rel="noopener noreferrer">{html.escape(oid)}</a>' if oid else '')
                          + (f' <span class="row-actions">'
                              f'<button class="copy-btn" data-order="{html.escape(oid)}" title="Copiar Order ID">📋</button>'
                              # Use buttons with data-order so JS can construct query params reliably
                              f'<button class="fill-btn" data-order="{html.escape(oid)}" title="Preencher formulário">↪️</button>'
                              f'<button class="open-detail" data-order="{html.escape(oid)}" title="Abrir detalhe">🔎</button>'
                              f'</span>')
                ) if oid else ''
            )

    # allow HTML (we will insert small markup for highlighting)
    html_table = df2.fillna('').to_html(index=False, table_id=table_id, classes='display', escape=False)

    # To avoid embedding external scripts inside the Streamlit iframe (which
    # triggers sandbox/feature warnings and may produce srcdoc syntax errors in
    # some browsers), we intentionally DO NOT include <script> tags or load
    # CDN JS here. The table remains styled and readable but without client-side
    # DataTables behavior. This keeps the console clean and the iframe safe.
    css_block = f"""
    <style>
    :root {{ --brand-900: #111922; --danger: #c62828; --gold: #caa85a; }}
    .neg {{ color: var(--danger); font-weight: 600; }} .rev {{ color: #2e7d32; font-weight: 700; }}
    /* basic table styling (no JS) */
    #{table_id} {{ border-collapse: collapse; width:100%; table-layout: auto; font-family: 'Segoe UI', Roboto, Arial, sans-serif; }}
    #{table_id} th, #{table_id} td {{ padding: 10px 8px; border-bottom: 1px solid #eee; vertical-align: middle; text-align: center; font-size:13px; color: #12232f; }}
    #{table_id} thead th {{ background: var(--brand-900); color: #fff; font-weight:700; }}
    #{table_id} tbody tr:nth-child(odd) td {{ background:#fbfbff; }}
    #{table_id} tbody tr:hover td {{ background: #f3f6f9; }}
     /* Order ID column: keep link on its own line and show the 3 action buttons below it
         without extra horizontal space. Make the link block-level and the actions a
         compact inline-flex row. Reduce padding so the cell fits tightly to content. */
     #{table_id} thead th:nth-child(2), #{table_id} td:nth-child(2) {{ max-width: 220px; white-space: normal; text-align: left; padding:6px 6px; }}
     #{table_id} td:nth-child(2) a {{ display: block; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 100%; }}
    /* small action button style (compact) */
    #{table_id} .copy-btn {{ background:var(--gold); color:var(--brand-900); border:none; padding:4px 6px; border-radius:6px; margin-right:4px; font-size:12px; }}
    #{table_id} .row-actions {{ display:inline-flex; gap:4px; margin-top:4px; vertical-align:middle; align-items:center; }}
    #{table_id} .row-actions .copy-btn, #{table_id} .row-actions .fill-btn, #{table_id} .row-actions .open-detail {{
        background: transparent; border: 1px solid rgba(0,0,0,0.06); padding:3px 6px; border-radius:6px; font-size:12px; cursor:pointer; text-decoration:none; color:var(--brand-900);
    }}
    #{table_id} .row-actions .copy-btn:hover, #{table_id} .row-actions .fill-btn:hover, #{table_id} .row-actions .open-detail:hover {{ background: rgba(0,0,0,0.04); }}
    .interactive-card {{ background: transparent; padding: 6px; }}
    /* Ensure any textarea or input auto-generated by pandas/streamlit inside
       our table remains visible: some runtimes/styles render these with
       transparent text or hidden borders. Force readable color/background. */
    /* textarea/input inside the generated table: make them clearly readable
       and interactive (for copy), with a visible white background and subtle
       border so text isn't hidden by overlays or inherited styles. */
    #{table_id} textarea, #{table_id} input {{
        color: var(--text) !important;
        background: #ffffff !important;
        border: 1px solid rgba(0,0,0,0.06) !important;
        box-shadow: none !important;
        resize: none !important;
        width: 100% !important;
        height: auto !important;
        padding: 4px 6px !important;
        box-sizing: border-box !important;
        font-family: inherit !important;
        font-size: 13px !important;
        line-height: 1.2 !important;
        pointer-events: auto !important;
        overflow: visible !important;
    }}
    /* Description cell: truncate long review descriptions visually with
       ellipsis, but keep full text in the title attribute for hover tooltip. */
    #{table_id} .desc-cell {{
        display: inline-block;
        max-width: 420px; /* reasonable max so table doesn't break */
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        vertical-align: middle;
        text-align: left;
        padding: 2px 4px;
    }}
    </style>
    """

    # Return only the styled table HTML (no scripts). This avoids sandbox escapes
    # and 'Unrecognized feature' warnings from the browser. If richer client-side
    # interactivity is required later, we should move to a supported Streamlit
    # component (ag-grid / st-aggrid) or serve a small separate static page.
    safe = css_block + f'<div class="interactive-card" style="max-height:520px; overflow:auto">{html_table}</div>'
    # Do NOT include inline <script> or iframe-embedded JS here. Returning
    # only CSS + safe HTML ensures Streamlit will render the table in the
    # main page without creating a sandboxed iframe that produces
    # 'Unrecognized feature' / sandbox-escape warnings in the browser.
    return safe

DB_PATH = Path('ml_devolucoes.db')


def _download_db_from_env():
    """If the local DB file is missing, try to download it from a URL in
    the environment variable SQLITE_REMOTE_URL. This allows the deployed app
    to fetch the same sqlite database you use locally without committing the
    binary into the git repo.
    """
    if DB_PATH.exists():
        return True

    url = os.environ.get('SQLITE_REMOTE_URL')
    if not url:
        # Nothing to do
        return False

    try:
        # Stream the download to avoid memory pressure
        resp = requests.get(url, stream=True, timeout=30)
        resp.raise_for_status()
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(DB_PATH, 'wb') as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except Exception as e:
        # If download fails, leave a small log file for debugging
        try:
            with open('db_download_error.txt', 'w', encoding='utf-8') as ef:
                ef.write(f"Failed to download {url}: {repr(e)}\n")
        except Exception:
            pass
        return False


# Attempt to fetch DB from remote when running in an environment where the
# repository doesn't contain the sqlite file (e.g. Streamlit Cloud). If the
# fetch fails, the rest of the app will surface an explanatory error later.
_download_db_from_env()

@st.cache_data
def get_months():
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT DISTINCT substr(data_venda,1,7) as ym FROM orders ORDER BY ym DESC", con)
    con.close()
    months = df['ym'].dropna().tolist()
    return months


@st.cache_data
def get_return_reasons():
    """Return a sorted list of distinct motivo_resultado values from orders/returns."""
    con = sqlite3.connect(DB_PATH)
    reasons = set()
    try:
        # check orders table first
        df = pd.read_sql('SELECT DISTINCT motivo_resultado FROM orders WHERE motivo_resultado IS NOT NULL', con)
        if not df.empty:
            reasons.update(df['motivo_resultado'].dropna().astype(str).tolist())
    except Exception:
        pass
    try:
        df2 = pd.read_sql('SELECT DISTINCT motivo_resultado FROM returns WHERE motivo_resultado IS NOT NULL', con)
        if not df2.empty:
            reasons.update(df2['motivo_resultado'].dropna().astype(str).tolist())
    except Exception:
        pass
    con.close()
    return sorted([r for r in reasons if r and str(r).strip()])

# config path for persisted list of motivos considered "passíveis"
CONFIG_DIR = Path('config')
CONFIG_DIR.mkdir(exist_ok=True)
PASSIVEIS_CONFIG = CONFIG_DIR / 'motivos_passiveis.json'

def load_saved_passiveis():
    try:
        if PASSIVEIS_CONFIG.exists():
            with open(PASSIVEIS_CONFIG, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    return [str(x) for x in data]
    except Exception:
        pass
    return []

def save_passiveis(lst):
    try:
        with open(PASSIVEIS_CONFIG, 'w', encoding='utf-8') as f:
            json.dump(lst, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

# Substrings used to identify reasons that are typically considered
# 'passíveis de reembolso' according to Mercado Livre guidance.
PASSIVEIS_MOTIVO_SUBSTRINGS = [
    # exact phrases and useful variants based on your screenshot
    'comprador comprou o produto errado',
    'comprou o produto errado',
    'encontrou um preço melhor',
    'preço melhor',
    'se arrependeu da compra',
    'se arrependeu',
    'arrependeu',
    'houve danos devido a problemas com a transportadora',
    'problemas com a transportadora',
    'transportadora',
    'mercado envios',
    'mercadoenvios'
]

@st.cache_data
def load_financials(month=None, month_from=None, month_to=None, only_pending=False, only_loss=False, sku_filter=None, motivo_filter=None):
    con = sqlite3.connect(DB_PATH)
    q = 'SELECT o.order_id, o.data_venda, o.total_brl, o._valor_passivel_extorno, o._valor_pendente, o.dinheiro_liberado, oi.sku, oi.preco_unitario, oi.unidades, o.resultado, o.mes_faturamento FROM orders o JOIN order_items oi ON o.order_id=oi.order_id'
    filters = []
    # support either a single month (backwards-compatible) or a month range
    if month:
        filters.append(f"substr(o.data_venda,1,7) = '{month}'")
    else:
        if month_from:
            filters.append(f"substr(o.data_venda,1,7) >= '{month_from}'")
        if month_to:
            filters.append(f"substr(o.data_venda,1,7) <= '{month_to}'")
    # filter to only rows that likely require an estorno: orders with negative total
    if only_loss:
        # ensure we return orders where the canonical total is negative
        filters.append('o.total_brl < 0')
    elif only_pending:
        # legacy heuristic: _valor_pendente > 0
        filters.append('o._valor_pendente > 0')
    if sku_filter:
        # simple like
        filters.append(f"oi.sku LIKE '%{sku_filter}%'")
    # motivo_filter can be a list of strings; match orders.motivo_resultado
    if motivo_filter:
        # sanitize values and build SQL IN clause
        vals = [str(v).replace("'", "''") for v in motivo_filter]
        if vals:
            quoted = ','.join([f"'{v}'" for v in vals])
            filters.append(f"o.motivo_resultado IN ({quoted})")
    if filters:
        q += ' WHERE ' + ' AND '.join(filters)
    # order losses first (more negative totals at the top) when using the loss filter,
    # otherwise fall back to pending heuristic ordering for debugging.
    if only_loss:
        q += ' ORDER BY o.total_brl ASC'
    else:
        q += ' ORDER BY o._valor_pendente DESC'
    df = pd.read_sql(q, con)
    con.close()
    # post-process types
    if 'data_venda' in df.columns:
        df['data_venda'] = pd.to_datetime(df['data_venda'], errors='coerce')
    numeric_cols = ['total_brl', '_valor_passivel_extorno', '_valor_pendente', 'dinheiro_liberado', 'preco_unitario']
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
    # Derived columns for clearer business semantics
    # 'prejuizo_real_signed' = signed total (negative when the order is a net loss)
    # 'prejuizo_real' = absolute magnitude of that prejudice (positive number)
    # 'prejuizo_pendente_calc' = magnitude still pending (>=0)
    # 'prejuizo_pendente_signed' = signed pending amount (negative when there is an outstanding loss)
    if 'total_brl' in df.columns:
        # "Prejuízo" should reflect the signed total reported by Mercado Livre
        # i.e. negative when the net result is a loss. Keep an explicit field
        # for UI convenience that is exactly the ledger `total_brl`.
        df['prejuizo_real_signed'] = df['total_brl']
        # magnitude-only (positive) when there is a loss, zero otherwise
        df['prejuizo_real'] = df['prejuizo_real_signed'].where(df['prejuizo_real_signed'] < 0, 0.0).abs()
    else:
        df['prejuizo_real_signed'] = 0.0
        df['prejuizo_real'] = 0.0
    if 'dinheiro_liberado' not in df.columns:
        df['dinheiro_liberado'] = 0.0
    # pending magnitude (positive) and signed version (negative to match ML UI semantics)
    # keep internal heuristics but they should NOT be presented to users as the
    # canonical "Prejuízo pendente". We'll keep the columns for debugging and
    # internal inspection but hide them from the main UI and exports.
    df['prejuizo_pendente_calc'] = (df['prejuizo_real'] - df['dinheiro_liberado']).clip(lower=0.0)
    df['prejuizo_pendente_signed'] = -df['prejuizo_pendente_calc']
    return df


def ensure_reviews_table():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS reviews (
            order_id TEXT PRIMARY KEY,
            reviewed INTEGER DEFAULT 0,
            reviewed_by TEXT,
            reviewed_at TEXT,
            review_description TEXT
        )
    ''')
    # Ensure older installations have the review_description column
    cur.execute("PRAGMA table_info(reviews)")
    cols = [r[1] for r in cur.fetchall()]
    if 'review_description' not in cols:
        try:
            cur.execute("ALTER TABLE reviews ADD COLUMN review_description TEXT")
        except Exception:
            pass
    con.commit()
    con.close()


def get_reviews_map():
    ensure_reviews_table()
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql('SELECT order_id, reviewed, reviewed_by, reviewed_at, review_description FROM reviews', con)
    con.close()
    if df.empty:
        return {}
    return df.set_index('order_id').to_dict(orient='index')


def set_review(order_id: str, reviewed: bool, user: str = 'operator', description: str = None):
    ensure_reviews_table()
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    # store timestamps in UTC to avoid server/local timezone drift
    now = datetime.now(tz=timezone.utc).isoformat()
    # Use REPLACE so we update existing rows; include review_description
    cur.execute('REPLACE INTO reviews (order_id, reviewed, reviewed_by, reviewed_at, review_description) VALUES (?,?,?,?,?)',
                (order_id, 1 if reviewed else 0, user if reviewed else None, now if reviewed else None, description if reviewed else None))
    try:
        con.commit()
    except Exception as e:
        # persist a small debug file to help diagnose write failures in prod
        try:
            with open('review_error.log', 'a', encoding='utf-8') as ef:
                ef.write(f"Commit failed for set_review order_id={order_id} reviewed={reviewed} error={repr(e)}\n")
        except Exception:
            pass
    finally:
        con.close()
    # Audit the action so we can trace whether reviews were attempted in prod
    try:
        save_action(order_id, user or 'operator', 'set_review', f'reviewed={1 if reviewed else 0} reviewed_at={now if reviewed else None}')
    except Exception:
        try:
            with open('review_error.log', 'a', encoding='utf-8') as ef:
                ef.write(f"save_action failed for order_id={order_id}\n")
        except Exception:
            pass


def create_xlsx_export(df: pd.DataFrame, path: Path, display_names: dict):
    # df already contains display columns and a 'Revisado' column
    try:
        # use context manager to ensure compatibility with newer pandas/openpyxl
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Export')
            wb = writer.book
            ws = writer.sheets['Export']
        # header style
        header_font = Font(bold=True, color='000000')
        for col_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            # column width
            max_len = max(df[col].astype(str).map(len).max() if not df.empty else 10, len(col))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
            # currency formatting if column looks like currency
            if col.lower().find('r$')!=-1 or 'valor' in col.lower() or 'receita' in col.lower() or 'preço' in col.lower() or 'preco' in col.lower():
                for r in range(2, 2 + len(df)):
                    try:
                        # use a safe number format (avoid escape sequence warning)
                        ws.cell(row=r, column=col_idx).number_format = '#,##0.00'
                    except Exception:
                        pass
            # If this column looks like the Order ID display name, add an Excel hyperlink for each cell
            try:
                if 'order id' in col.lower() or col.lower().strip() == 'order':
                    for r_idx, val in enumerate(df[col].astype(str), start=2):
                        try:
                            if val and val.strip():
                                url = f'https://www.mercadolivre.com.br/vendas/{val}/detalhe'
                                c = ws.cell(row=r_idx, column=col_idx)
                                c.value = val
                                c.hyperlink = url
                                c.font = Font(color='0000EE', underline='single')
                                c.alignment = Alignment(horizontal='center')
                        except Exception:
                            pass
            except Exception:
                pass
        return True, None
    except Exception as e:
        return False, str(e)

def save_action(order_id: str, user: str, action: str, note: str):
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    # create actions table if it doesn't exist (simple audit table)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS actions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id TEXT,
            user TEXT,
            action TEXT,
            note TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    ''')
    cur.execute('INSERT INTO actions (order_id, user, action, note) VALUES (?,?,?,?)', (order_id, user, action, note))
    con.commit()
    con.close()

def main():
    # page config (favicon will be set after assets are resolved below)
    # --- DEBUG: quick visual check for header/logo rendering ---
    # show the primary asset inline only if the file actually exists to avoid
    # Streamlit registering a media id that may be missing in the runtime.
    try:
        img_path = Path('assets') / '2.png'
        if img_path.exists() and img_path.stat().st_size > 0:
            st.image(str(img_path), width=120)
    except Exception:
        # ignore any failure displaying inline images
        pass
    # brand/theme CSS polish (colors, spacing, table and buttons)
    st.markdown(f"""
    <style>
    :root {{
        --brand-900: #0b2f44; /* darker, high-contrast navy */
        --brand-800: #0f3b53;
        --brand-600: #1b5670;
        --brand-500: #2b6f8a;
        --gold: #caa85a;
        --bg: #f6f9fb;
        --muted: #6b7280;
        --danger: #c62828;
        --text: #12232f;
    }}

    /* Page background and readable base color */
    .stApp {{
        background: linear-gradient(180deg, var(--bg), #ffffff);
        padding-top: 0.8rem !important;
        color: var(--text);
        -webkit-font-smoothing: antialiased;
    }}
    .block-container, .stApp .block-container {{ padding: 0.8rem 1.4rem !important; max-width: 1280px !important; margin: 0 auto !important; }}

    /* Topbar (navy) to host logo and title - ensure strong contrast */
    .topbar {{ background: var(--brand-900); color: #fff; padding: 12px 22px; border-radius: 8px; margin-bottom: 18px; box-shadow: 0 8px 20px rgba(11,47,68,0.12); position: relative; z-index:2; width: 100%; }}
    .brand-header {{ display:flex; align-items:center; gap:18px; margin:0; }}
    .brand-logo img.brand-logo {{ width:120px; height:120px; object-fit:contain; border-radius:10px; box-shadow: 0 8px 24px rgba(11,47,68,0.12); }}
    .brand-logo {{ width:120px; height:120px; flex:0 0 120px; display:flex; align-items:center; justify-content:center; }}
    .brand-text h1 {{ margin:0; font-size:28px; color: #fff; font-weight:800; line-height:1.02; text-shadow: 0 2px 6px rgba(0,0,0,0.25); }}
    .brand-text .brand-sub {{ margin-top:6px; color: rgba(255,255,255,0.9); font-size:14px; }}
    /* Prevent selection of header text and make logo visually above the title */
    .brand-text, .brand-text h1, .brand-text .brand-sub {{
        -webkit-user-select: none; /* Safari */
        -moz-user-select: none; /* Firefox */
        -ms-user-select: none; /* IE10+ */
        user-select: none;
    }}
    .brand-logo img.brand-logo {{ position: relative; z-index: 5; }}
    .brand-text h1 {{ position: relative; z-index: 1; padding-top: 6px; }}
    .brand-decor {{ position:absolute; right:22px; top:12px; width:96px; height:96px; opacity:0.95; filter: drop-shadow(0 6px 18px rgba(0,0,0,0.12)); border-radius:8px; max-width:96px; max-height:96px; object-fit:contain; }}

    /* Card surface for main panels: use white surfaces to improve contrast */
    .main-card, .interactive-card, .stApp .block-container > :where(div) {{ background: #ffffff; border-radius: 8px; border:1px solid rgba(11,47,68,0.04); box-shadow: 0 6px 18px rgba(11,47,68,0.04); padding: 16px; }}

    /* Sidebar: subtle, but keep good contrast for text */
    .stSidebar {{ background: linear-gradient(180deg,#f7fafc,#f1f5f8) !important; box-shadow: inset -4px 0 18px rgba(11,47,68,0.02) !important; color: var(--text) !important; }}

    /* Buttons: clearer CTA style using brand colors */
    .stButton button, .stButton>button, .stDownloadButton>button {{ background: linear-gradient(180deg,var(--brand-600),var(--brand-800)) !important; color: #fff !important; border: 1px solid rgba(0,0,0,0.06) !important; padding: 10px 14px !important; border-radius:10px !important; box-shadow: 0 8px 20px rgba(11,47,68,0.06) !important; transition: transform .12s ease, box-shadow .12s ease, background .12s ease; }}
    .stButton button:hover, .stDownloadButton>button:hover {{ transform: translateY(-2px); box-shadow: 0 18px 44px rgba(11,47,68,0.12) !important; }}

    /* Links and accents */
    a, a:hover, .stApp a {{ color: #0b66a3 !important; text-decoration: none !important; }}

    /* DataTables global tweaks: ensure headers and cells are readable */
    table.display thead th, table.dataframe thead th, table.display thead td {{ background: var(--brand-900) !important; color: #fff !important; border-bottom: 1px solid rgba(0,0,0,0.06) !important; }}
    table.display tbody td, table.dataframe tbody td {{ background: #ffffff !important; color: var(--text) !important; }}
    table.display, table.dataframe {{ width: 100% !important; border-collapse: collapse !important; font-family: 'Segoe UI', Roboto, Arial, sans-serif !important; font-size:13px !important; }}

    /* inline action button style */
    #sample_tbl .copy-btn {{ background: var(--gold) !important; color: var(--brand-900) !important; border:none !important; padding:6px 8px !important; border-radius:6px !important; cursor:pointer !important; margin-right:6px !important; box-shadow: 0 6px 18px rgba(11,47,68,0.06) !important; }}
    #sample_tbl .copy-btn:hover {{ background: #e6c889 !important; transform: translateY(-2px) !important; }}
    #sample_tbl .open-link {{ color: var(--brand-500) !important; text-decoration: none !important; font-weight:600 !important; }}

    /* Misc */
    .neg {{ color: var(--danger) !important; font-weight: 700 !important; }}
    .rev {{ color: #2e7d32 !important; font-weight: 700 !important; }}

    @media (max-width: 1000px) {{
        .block-container, .stApp .block-container {{ padding-left: 0.75rem !important; padding-right: 0.75rem !important; }}
        table.display {{ font-size: 12px !important; }}
        .brand-logo img.brand-logo {{ width:84px; height:84px; }}
        .brand-decor {{ display:none; }}
        .topbar {{ width: 100%; margin-left:0; padding-right: 24px; border-radius: 8px; }}
    }}

    </style>
    """, unsafe_allow_html=True)

    # ensure assets directory exists for user-supplied logos
    assets_dir = Path('assets')
    assets_dir.mkdir(exist_ok=True)

    # asset paths (defaults inside workspace)
    logo_center_path = assets_dir / 'logo_center.png'
    alt_logo2 = assets_dir / '2.png'
    logo_round_path = assets_dir / 'logo_round.png'
    favicon_path = assets_dir / 'favicon.png'
    favicon_gold = assets_dir / 'favicon_gold.png'

    # determine page favicon (prefer round SVG as favicon-like icon)
    # NOTE: avoid passing a missing/invalid media id to Streamlit which can
    # raise MediaFileStorageError in some deployment environments. Use a
    # small emoji fallback when no workspace asset is present.
    page_favicon = None
    try:
        if logo_round_path.exists():
            page_favicon = str(logo_round_path)
        elif favicon_path.exists():
            page_favicon = str(favicon_path)
        elif favicon_gold.exists():
            page_favicon = str(favicon_gold)
        else:
            # simple emoji is safe and won't trigger media file lookups
            page_favicon = "📊"
    except Exception:
        page_favicon = "📊"
    st.set_page_config(page_title='BI Devoluções - Protótipo', layout='wide', page_icon=page_favicon)

    # Fallback: look for SVG logos in the user's design folder if not present in assets
    try:
        external_dir = Path.home() / 'Documents' / 'design' / 'screenshots'
        if external_dir.exists() and external_dir.is_dir():
            svgs = list(external_dir.glob('*.svg'))
            if svgs:
                # prefer names containing 'nome' or 'logo' for center, 'simbol' or 'simbolo' for round
                center_svg = next((s for s in svgs if 'nome' in s.name.lower() or 'logo' in s.name.lower()), None)
                round_svg = next((s for s in svgs if 'simbol' in s.name.lower() or 'symbol' in s.name.lower()), None)
                # fallback to first svg if specific patterns not found
                if not center_svg:
                    center_svg = svgs[0]
                if not round_svg:
                    round_svg = svgs[0] if svgs else None
                # assign paths (they may be outside workspace but Streamlit can load absolute paths)
                if center_svg:
                    logo_center_path = center_svg
                if round_svg:
                    logo_round_path = round_svg
    except Exception:
        # best-effort: if anything fails, continue using workspace assets
        pass

    # Prefer logos that live inside the workspace assets folder (these are served by Streamlit).
    # Look for common asset names (SVG preferred). If only an external SVG was found above,
    # copy it into assets so the browser can load it via /assets/<name>.
    import shutil
    logo_asset = None
    for candidate in ['logo_center_v2.svg', 'logo_center.svg', 'logo_center.png', '2.png', 'logo_round.svg', 'logo_round.png']:
        p = assets_dir / candidate
        if p.exists():
            logo_asset = p
            break

    # If we didn't find a logo in assets but have an external logo_center_path, copy it into assets
    if logo_asset is None and logo_center_path and Path(logo_center_path).exists():
        try:
            ext = Path(logo_center_path)
            dest = assets_dir / 'logo_center.svg' if ext.suffix.lower()=='.svg' else assets_dir / ('logo_center' + ext.suffix)
            shutil.copy2(ext, dest)
            logo_asset = dest
        except Exception:
            logo_asset = None


    # prefer an explicit favicon if present
    try:
        if favicon_path.exists() and favicon_path.stat().st_size > 0:
            st.markdown(f"<link rel=\"icon\" href=\"/assets/{favicon_path.name}\">", unsafe_allow_html=True)
        elif favicon_gold.exists() and favicon_gold.stat().st_size > 0:
            st.markdown(f"<link rel=\"icon\" href=\"/assets/{favicon_gold.name}\">", unsafe_allow_html=True)
        else:
            # no-op: rely on page_icon passed to set_page_config (emoji fallback)
            pass
    except Exception:
        # If anything goes wrong when trying to reference assets, avoid raising
        # so the app can continue to run. The page_icon emoji is sufficient.
        pass

    # Header: use a single flex HTML block so we can precisely position logo and centered title
    # Prefer serving the workspace asset 'logo_center_v2.svg' if present
    # Inline SVG content for robust rendering (prefer v2 then v1)
    # Simpler, robust header: prefer a workspace-served asset and always build the header HTML
    logo_name = None
    # prefer non-empty assets; some SVG copies may be empty/corrupt — require >100 bytes
    for candidate in ('logo_center_v2.svg', 'logo_center.svg', 'logo_center.png', '2.png'):
        p = assets_dir / candidate
        try:
            if p.exists() and p.stat().st_size > 100:
                logo_name = candidate
                break
        except Exception:
            continue

    if logo_name:
        # Read the file and embed as data URI to avoid any serving/caching issues
        p = assets_dir / logo_name
        try:
            data = p.read_bytes()
            b64 = base64.b64encode(data).decode('ascii')
            if p.suffix.lower() == '.svg':
                src = f"data:image/svg+xml;base64,{b64}"
            else:
                # assume png for other extensions
                src = f"data:image/png;base64,{b64}"
            # restore larger logo width requested by user and keep it inside the content flow
            logo_tag = f"<img src='{src}' style='width:320px;height:auto;display:block;' alt='logo'>"
        except Exception:
            logo_tag = "<div style='width:220px;height:1px;'></div>"
    else:
        # fallback spacer so the title stays centered even without an image
        logo_tag = "<div style='width:360px;height:1px;'></div>"

    # Build a white content card and use CSS grid so the logo appears in the
    # top-left of the card and the title is centered beneath it (spanning the
    # full card width). This matches the user's mock where the title is below
    # the logo, not beside it.
    header_html = (
        "<div style='width:100%; display:flex; justify-content:center; padding:18px 6px;'>"
        "<div style='max-width:1100px; width:100%; background:#ffffff; border-radius:8px; padding:18px; box-shadow: 0 6px 18px rgba(11,47,68,0.04);'>"
        "  <div style='display:grid; grid-template-columns: 320px 1fr; grid-template-rows: auto auto; gap:12px; align-items:start;'>"
        "    <div style='grid-column:1; grid-row:1; display:block; align-self:start;'>"
        f"      {logo_tag}"
        "    </div>"
        "    <div style='grid-column:2; grid-row:1;'></div>"  # empty cell to the right of logo
        "    <div style='grid-column:1 / -1; grid-row:2; text-align:center; padding-top:6px;'>"
        "      <h1 style='margin:0;color:var(--brand-900);font-weight:800;font-size:32px;line-height:1.04;'>Controle de Devoluções — Protótipo</h1>"
        "      <div style='color:rgba(17,25,34,0.7);margin-top:6px;font-size:14px;'>Painel para triagem, revisão e exportação de casos de devolução</div>"
        "    </div>"
        "  </div>"
        "</div></div>"
    )
    st.markdown(header_html, unsafe_allow_html=True)

    # NOTE: removed parent-side postMessage listener to avoid injecting
    # runtime JS into the main page. All table actions are rendered as
    # safe links or handled server-side; keeping the markup JS-free avoids
    # sandbox warnings and cross-origin issues in deployed environments.

    months = get_months()
    col1, col2, col3 = st.columns([2,2,1])
    with col1:
        # month range filter: allow selecting start and end months (YYYY-MM). Empty = no bound
        month_from = st.selectbox('Mês início (YYYY‑MM)', options=[''] + months, index=0, help='Mês inicial do período (vazio = sem limite inferior)')
        month_to = st.selectbox('Mês fim (YYYY‑MM)', options=[''] + months, index=0, help='Mês final do período (vazio = sem limite superior)')
    with col2:
        sku = st.text_input('Filtro SKU (parte)', help='Filtre por parte do SKU (case-insensitive, substring).')
    # reasons filter (populate from DB)
    reasons = sorted(set(get_return_reasons()))
    motivos_selected = []
    auto_filter = False
    # load saved passíveis (config)
    saved_passiveis = load_saved_passiveis()
    if reasons:
        with col2:
            # checkbox to auto-apply common passible motivos (auto heuristics)
            auto_filter = st.checkbox('Somente motivos passíveis de reembolso (auto)', value=False, help='Aplicar filtro automático com motivos tipicamente passíveis de reembolso')
            if auto_filter:
                # find reasons that match any of the substrings
                lowered = [(r, r.lower()) for r in reasons]
                auto_sel = [r for r, rl in lowered if any(sub in rl for sub in PASSIVEIS_MOTIVO_SUBSTRINGS)]
                # augment with saved_passiveis if present
                merged = sorted(set(auto_sel + [s for s in saved_passiveis if s in reasons]))
                if not merged:
                    st.warning('Nenhum motivo automático encontrado — revise a lista de motivos disponíveis ou edite o mapeamento.')
                motivos_selected = st.multiselect('Motivo da devolução (filtrar)', options=reasons, default=merged, help='Selecione um ou mais motivos para filtrar as devoluções retornadas pela plataforma')
                if merged:
                    st.caption(f'Filtro automático aplicou {len(merged)} motivos (heurística + salvos).')
            else:
                motivos_selected = st.multiselect('Motivo da devolução (filtrar)', options=reasons, help='Selecione um ou mais motivos para filtrar as devoluções retornadas pela plataforma')

        # management expander: allow saving the definitive list of passive motivos
        with st.expander('Gerenciar motivos passíveis (salvos)', expanded=False):
            st.write('Motivos detectados na base:')
            st.write(f'{len(reasons)} motivos encontrados.')
            # show a multiselect prefilled with saved_passiveis so the user can edit and save
            new_saved = st.multiselect('Marque motivos que considera passíveis', options=reasons, default=[s for s in saved_passiveis if s in reasons])
            if st.button('Salvar motivos passíveis'):
                ok = save_passiveis(new_saved)
                if ok:
                    st.success('Lista salva em config/motivos_passiveis.json')
                else:
                    st.error('Falha ao salvar lista — verifique permissões de arquivo.')
    with col3:
        only_pending = st.checkbox('Apenas pendentes (heurística antiga)', value=False, help='Heurística interna: mostra pedidos com _valor_pendente > 0 (útil para debugging).')
        only_loss = st.checkbox('Apenas com prejuízo (Total < 0)', value=True, help='Filtrar apenas pedidos com total negativo — estes são passíveis de estorno e são o foco desta tela.')

    # load reviews map (we will fetch fresh copies right before rendering
    # UI blocks that depend on reviews so recent writes via set_review are
    # immediately visible without requiring a full page reload).
    # Note: some export paths fetch the map on-demand; here we deliberately
    # avoid assigning to a long-lived variable that can become stale.

    # df: the dataset used by the Returns tab (respects only_loss)
    # prefer explicit month range; pass through the month_from/month_to values
    mf = month_from if month_from else None
    mt = month_to if month_to else None
    df = load_financials(month=None, month_from=mf, month_to=mt, only_pending=only_pending, only_loss=only_loss, sku_filter=sku if sku else None, motivo_filter=motivos_selected if motivos_selected else None)
    # df_all: full dataset for the selected filters except the only_loss filter — used for Metrics
    df_all = load_financials(month=None, month_from=mf, month_to=mt, only_pending=False, only_loss=False, sku_filter=sku if sku else None, motivo_filter=motivos_selected if motivos_selected else None)

    # helper: format currency BRL
    def fmt_brl(v):
        try:
            return f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except Exception:
            return str(v)

    def fmt_brl_signed(v):
        try:
            fv = float(v)
        except Exception:
            return str(v)
        if fv < 0:
            return f"-R$ {abs(fv):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"R$ {fv:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    # Split UI into two tabs: Metrics and Returns (Devoluções)
    tab_metrics, tab_returns = st.tabs(['Métricas', 'Devoluções'])

    with tab_metrics:
        st.subheader('Métricas')
        st.write('KPIs e métricas gerais sobre o período/filtro atual')
        # KPIs
        k1, k2, k3, k4 = st.columns(4)
        # Use df_all for metrics so KPIs reflect the full selection (not only the loss-filtered view)
        total_revenue = df_all['total_brl'].sum() if 'total_brl' in df_all.columns else 0.0
        total_orders = df_all['order_id'].nunique() if 'order_id' in df_all.columns else 0
        # pending prejudice from the loss-filtered df (what is outstanding)
        total_pending = df['prejuizo_pendente_signed'].sum() if 'prejuizo_pendente_signed' in df.columns else (-df['_valor_pendente'].sum() if '_valor_pendente' in df.columns else 0.0)
        top_skus = df_all.groupby('sku').agg(revenue=('total_brl','sum')).sort_values('revenue', ascending=False).head(10)

        k1.metric('Receita (seleção)', fmt_brl(total_revenue))
        k2.metric('Pedidos (seleção)', f'{total_orders:,}')
        k3.metric('Prejuízo pendente (visão devoluções)', fmt_brl_signed(total_pending))
        k4.metric('Top SKU (receita)', f"{top_skus.index[0] if not top_skus.empty else '-'}")

    # (Resumo de devoluções removido daqui — agora exibido na aba Devoluções para evitar redundância)

        st.markdown('---')
        st.subheader('Curva ABC (por receita) — Top SKUs')
        if not top_skus.empty:
            # use matplotlib to avoid Altair/vega incompatibilities in this environment
            fig, ax = plt.subplots(figsize=(8, 3))
            ax.bar(top_skus.index.astype(str), top_skus['revenue'])
            ax.set_ylabel('Receita')
            ax.set_xlabel('SKU')
            # set explicit tick positions before labeling to avoid matplotlib UserWarning
            ax.set_xticks(range(len(top_skus.index)))
            ax.set_xticklabels(top_skus.index.astype(str), rotation=45, ha='right')
            st.pyplot(fig)
        else:
            st.info('Sem dados para curva ABC com o filtro atual.')

        st.markdown('---')
        st.subheader('Ticket médio (por pedido)')
        if total_orders > 0:
            ticket_medio = total_revenue / total_orders
            st.write('Ticket médio da seleção: ', fmt_brl(ticket_medio))
        else:
            st.info('Nenhum pedido na seleção para calcular ticket médio.')

        st.markdown('---')
        st.subheader('Taxa de devolução e evolução diária')
        # return rate: fraction of orders with total_brl < 0
        if total_orders > 0:
            returns_count = df_all['total_brl'].lt(0).sum()
            return_rate = returns_count / total_orders
            st.metric('Taxa de devolução (pedidos)', f"{return_rate:.2%}", f"{returns_count} pedidos")
        else:
            st.info('Sem pedidos na seleção para calcular taxa de devolução.')

        # daily evolution (revenue and returns)
        if 'data_venda' in df_all.columns and not df_all.empty:
            daily = df_all.copy()
            # normalize to midnight datetimes for resampling convenience
            daily['date'] = pd.to_datetime(daily['data_venda']).dt.normalize()
            daily_agg = daily.groupby('date').agg(total_revenue=('total_brl','sum'), orders=('order_id','nunique'), returns_count=('total_brl', lambda s: (s<0).sum())).reset_index()

            # Choose aggregation level when the time series is long to avoid label overdraw.
            n_points = len(daily_agg)
            if n_points > 365:
                # long range -> monthly
                plot_df = daily_agg.set_index('date').resample('M').sum().reset_index()
                date_locator = mdates.AutoDateLocator(minticks=4, maxticks=8)
                date_formatter = mdates.AutoDateFormatter(date_locator)
            elif n_points > 90:
                # medium range -> weekly
                plot_df = daily_agg.set_index('date').resample('W').sum().reset_index()
                date_locator = mdates.AutoDateLocator(minticks=6, maxticks=12)
                date_formatter = mdates.AutoDateFormatter(date_locator)
            else:
                # short range -> daily
                plot_df = daily_agg.copy()
                date_locator = mdates.AutoDateLocator()
                date_formatter = mdates.AutoDateFormatter(date_locator)

            # matplotlib line + bar using proper date axis (avoids messy string ticks)
            fig2, (ax1, ax2) = plt.subplots(nrows=2, ncols=1, figsize=(10,6), sharex=True)
            ax1.plot(plot_df['date'], plot_df['total_revenue'], marker='o')
            ax1.set_ylabel('Receita')
            ax2.bar(plot_df['date'], plot_df['returns_count'])
            ax2.set_ylabel('Devoluções (count)')

            # format x-axis with date locator/formatter and readable rotation
            ax2.xaxis.set_major_locator(date_locator)
            ax2.xaxis.set_major_formatter(date_formatter)
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
            fig2.tight_layout()
            st.pyplot(fig2)
        else:
            st.info('Sem série temporal para mostrar evolução diária.')

    with tab_returns:
        st.subheader('Lista de Pedidos (amostra)')
        st.write('Tabela interativa — revise os pedidos e marque como "Revisado" quando concluído.')
        # KPIs specific to the Returns view (reflect the active df filters)
        try:
            table_rows = len(df)
            table_orders = df['order_id'].nunique() if 'order_id' in df.columns else 0
            table_sum_prejuizo = df['prejuizo_real_signed'].sum() if 'prejuizo_real_signed' in df.columns else 0.0
            table_sum_pendente = df['prejuizo_pendente_calc'].sum() if 'prejuizo_pendente_calc' in df.columns else 0.0
            rc1, rc2 = st.columns(2)
            rc1.metric('Pedidos retornados', f'{table_orders:,}')
            rc2.metric('Soma Prejuízo', fmt_brl_signed(table_sum_prejuizo))
            st.markdown('---')
        except Exception:
            # gracefully ignore if df missing/invalid
            pass
        # prefer interactive dataframe only if pyarrow is available (st.dataframe imports pyarrow)
        sample = df.head(200).copy()
        # nice formatting for sample
        if 'data_venda' in sample.columns:
            sample['data_venda'] = sample['data_venda'].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
        for col in ['total_brl', '_valor_passivel_extorno', '_valor_pendente', 'dinheiro_liberado', 'preco_unitario']:
            if col in sample.columns:
                sample[col] = sample[col].apply(lambda x: fmt_brl(x) if pd.notna(x) else '')

        # integrate 'Revisado' status from reviews table and prepare display-only DF
        display_map_ui = {
            'order_id':'Order ID',
            'data_venda':'Data da venda',
            'total_brl':'Total (R$)',
            '_valor_passivel_extorno':'Passível de estorno (R$)',
            'prejuizo_real_signed':'Prejuízo (R$)',
            'dinheiro_liberado':'Dinheiro liberado (R$)',
            'sku':'SKU',
            'preco_unitario':'Preço unitário (R$)',
            'unidades':'Unidades',
            'resultado':'Resultado'
        }

        if not sample.empty:
            sample_display = sample.copy()
            # Fetch the latest reviews map here so that any set_review calls
            # performed earlier in the same session (or just before rendering)
            # are reflected in the table immediately.
            reviews_map = get_reviews_map()
            sample_display['Revisado'] = sample_display['order_id'].apply(lambda oid: '✅' if bool(reviews_map.get(str(oid), {}).get('reviewed', 0)) else '')
            sample_display['Revisado_por'] = sample_display['order_id'].apply(lambda oid: reviews_map.get(str(oid), {}).get('reviewed_by'))
            # assign raw value (may be ISO string or None)
            sample_display['Revisado_em'] = sample_display['order_id'].apply(lambda oid: reviews_map.get(str(oid), {}).get('reviewed_at'))
            # bring the review textual description from reviews_map into the
            # display DF so the 'Descrição' column in the table shows what was
            # saved in the review form (matches the detail view behavior).
            try:
                sample_display['review_description'] = sample_display['order_id'].apply(lambda oid: reviews_map.get(str(oid), {}).get('review_description') or '')
            except Exception:
                sample_display['review_description'] = ''
            # Convert the Revisado_em values to the same display format used in
            # the detailed view: parse the ISO/naive values and convert to
            # America/Sao_Paulo (UTC-3) using the centralized converter. This
            # ensures list and detail show the same local time.
            try:
                tmp = pd.DataFrame({'Revisado_em': sample_display['Revisado_em']})
                tmp = _convert_ts_for_display(tmp, ts_cols='Revisado_em')
                sample_display['Revisado_em'] = tmp['Revisado_em'].fillna('').astype(str)
            except Exception:
                # fallback: ensure string type so the table doesn't break
                sample_display['Revisado_em'] = sample_display['Revisado_em'].apply(lambda v: str(v) if pd.notna(v) else '')
            # format prejuízo: we now use signed columns so negatives are shown (ML UI shows negative values)
            # prejuizo_real_signed and prejuizo_pendente_signed contain negative numbers when there's a loss
            if 'prejuizo_real_signed' in sample_display.columns:
                def fmt_signed_currency(x):
                    try:
                        v = float(x)
                    except Exception:
                        return x
                    s = f"R$ {abs(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    if v < 0:
                        return f"<span class='neg'>-{s}</span>"
                    return s
                sample_display['prejuizo_real_signed'] = sample_display['prejuizo_real_signed'].apply(fmt_signed_currency)
            if 'prejuizo_pendente_signed' in sample_display.columns:
                sample_display['prejuizo_pendente_signed'] = sample_display['prejuizo_pendente_signed'].apply(lambda v: f"<span class='neg'>-R$ {float(abs(v)):,.2f}</span>" if float(v) < 0 else f"R$ {float(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
            # Build a user-friendly view with desired column order and names
            # Desired order: Order ID, Data da venda, Preço unitário, Total (use _valor_passivel_extorno), Prejuízo, SKU, Unidades, Resultado, Revisado metadata
            view_cols = []
            # Order ID
            view_cols.append('order_id')
            # Data da venda
            if 'data_venda' in sample_display.columns:
                view_cols.append('data_venda')
            # Preço unitário
            if 'preco_unitario' in sample_display.columns:
                view_cols.append('preco_unitario')
            # Use _valor_passivel_extorno as the Total column (rename later)
            if '_valor_passivel_extorno' in sample_display.columns:
                view_cols.append('_valor_passivel_extorno')
            # Prejuízo (signed)
            if 'prejuizo_real_signed' in sample_display.columns:
                view_cols.append('prejuizo_real_signed')
            # SKU, Unidades
            for c in ['sku', 'unidades']:
                if c in sample_display.columns:
                    view_cols.append(c)

            # append review metadata
            for extra in ['Revisado', 'Revisado_por', 'Revisado_em']:
                if extra in sample_display.columns:
                    view_cols.append(extra)
            # add the textual description (use review_description from reviews table if present)
            # we'll map it to the 'Descrição' column later and place it last
            # Build the display-friendly 'Descrição' column. Prefer the
            # review_description stored in the reviews table; if absent,
            # fall back to the resultado column. To avoid breaking the
            # table layout when descriptions are long, truncate the
            # visible text to 100 characters and render an HTML element
            # with a tooltip (title) that contains the full text.
            def _make_desc_cell(val, limit=100):
                try:
                    s = '' if val is None else str(val)
                except Exception:
                    s = ''
                full_esc = html.escape(s)
                if len(s) > limit:
                    short = html.escape(s[:limit].rstrip()) + '...'
                else:
                    short = full_esc
                # use a div with class desc-cell so CSS can ellipsize it
                return f"<div class='desc-cell' title=\"{full_esc}\">{short}</div>"

            if 'review_description' in sample_display.columns:
                sample_display['Descrição'] = sample_display['review_description'].apply(lambda v: _make_desc_cell(v, limit=100))
            elif 'resultado' in sample_display.columns:
                # fall back to existing resultado column if no review_description
                sample_display['Descrição'] = sample_display['resultado'].apply(lambda v: _make_desc_cell(v, limit=100))
            # ensure 'Descrição' is the last column
            if 'Descrição' in sample_display.columns:
                view_cols.append('Descrição')

            sample_display = sample_display[view_cols].copy()
            # add a 1-based index column for easier reference in the UI
            sample_display.insert(0, '#', range(1, 1 + len(sample_display)))

            # (debug output removed) — avoid showing internal debug table above
            # the main sample table in production. Use SHOW_REVIEW_DEBUG only
            # in specific troubleshooting builds if needed.

            # format numeric columns for display
            if 'preco_unitario' in sample_display.columns:
                sample_display['preco_unitario'] = sample_display['preco_unitario'].apply(lambda x: fmt_brl(x) if pd.notna(x) else '')
            if '_valor_passivel_extorno' in sample_display.columns:
                sample_display['_valor_passivel_extorno'] = sample_display['_valor_passivel_extorno'].apply(lambda x: fmt_brl(x) if pd.notna(x) else '')
            if 'prejuizo_real_signed' in sample_display.columns:
                sample_display['prejuizo_real_signed'] = sample_display['prejuizo_real_signed'].apply(lambda x: fmt_brl_signed(x) if pd.notna(x) else '')

            # rename columns to friendly Portuguese names
            rename_map = {
                'order_id':'Order ID',
                'data_venda':'Data da venda',
                'preco_unitario':'Preço unitário (R$)',
                '_valor_passivel_extorno':'Total (R$)',
                'prejuizo_real_signed':'Prejuízo (R$)',
                'sku':'SKU',
                'unidades':'Unidades',
                # 'resultado' removed — using 'Descrição' instead for review notes
                'Descrição':'Descrição'
            }
            sample_display.rename(columns=rename_map, inplace=True)
        else:
            sample_display = sample

        # render interactive table using client-side DataTables (no pyarrow required)
        # Render the table as inline HTML (no components iframe). Avoids
        # creating a sandboxed iframe that triggers browser 'Unrecognized
        # feature' warnings or allows iframe escapes in some hosting
        # environments. The table is static HTML+CSS (no JS actions).
        try:
            html_snippet = render_interactive_table(sample_display, table_id='sample_tbl')
            st.markdown(html_snippet, unsafe_allow_html=True)
        except Exception:
            st.markdown(render_interactive_table(sample_display, table_id='sample_tbl'), unsafe_allow_html=True)

        # Copy-visible-order-ids helper: collect the Order IDs currently shown
        # in the sample_display and offer a one-click copy button. If the
        # clipboard API isn't available, fall back to showing a textarea with
        # the IDs so the user can copy manually.
        try:
            visible_oids = []
            # sample_display may have been renamed to display-friendly headings
            if 'Order ID' in sample_display.columns:
                visible_oids = sample_display['Order ID'].astype(str).tolist()
            else:
                visible_oids = sample_display['order_id'].astype(str).tolist()
            oids_text = '\n'.join([o for o in visible_oids if o and o.strip()])
        except Exception:
            oids_text = ''

        if oids_text:
            copy_col1, copy_col2 = st.columns([1,3])
            with copy_col1:
                if st.button('Copiar Order IDs visíveis'):
                    # small JS snippet to copy to clipboard. If the environment
                    # blocks clipboard access, show the textarea below instead.
                    # Build a JSON-escaped JS string to safely embed newlines/quotes
                    # Copy helper: rather than injecting JS into the page, show
                    # the visible Order IDs inside a textarea so the user can
                    # copy manually. This avoids clipboard API calls that may be
                    # blocked in some hosting environments and removes the need
                    # for components-based JS injection.
                    st.text_area('Order IDs (copie manualmente)', value=oids_text, height=120)
            with copy_col2:
                st.caption('IDs mostrados na tabela acima. Use o botão para copiar todos os Order IDs exibidos (um por linha).')

        # Read query params so action buttons that navigate to ?prefill_order_id=... or
        # ?detail_id=... can prefill the review form or open details without copy/paste.
        # Read query params using the supported API. Fall back to the
        # experimental method if `st.query_params` is not available in
        # very old Streamlit runtimes. Normalize access so that repeated
        # keys are handled correctly: prefer get_all() when available.
        try:
            _qp = st.query_params
        except Exception:
            try:
                _qp = st.experimental_get_query_params()
            except Exception:
                _qp = {}

        def _qp_first(key):
            """Return the first value for a query param key or None.

            Works with both the modern `st.query_params` (which provides
            get_all) and the legacy dict-of-lists returned by
            `st.experimental_get_query_params()`.
            """
            try:
                if hasattr(_qp, 'get_all'):
                    vals = _qp.get_all(key)
                    return vals[0] if vals else None
                # legacy: mapping to list-of-values
                v = _qp.get(key)
                if isinstance(v, list):
                    return v[0] if v else None
                return v
            except Exception:
                return None

        prefill_order = _qp_first('prefill_order_id')
        detail_prefill = _qp_first('detail_id')

        st.subheader('Marcar/Revisar pedido')
        with st.form('review_form'):
            order_id_in = st.text_input('Order ID para marcar/revisar', value=str(prefill_order) if prefill_order else '')
            user_in = st.text_input('Usuário (quem revisou)', value='operator')
            reviewed_in = st.checkbox('Marcar como Revisado', value=False)
            review_description_in = st.text_area('Descrição / Resumo da contestação (visível no campo Descrição)', height=120)
            submitted = st.form_submit_button('Salvar revisão')
            if submitted:
                if not order_id_in:
                    st.error('Informe um Order ID')
                else:
                    set_review(order_id_in, reviewed_in, user_in, review_description_in)
                    st.success(('Marcado como Revisado' if reviewed_in else 'Desmarcado revisão') + ' para order_id=' + order_id_in)
                    # Force a rerun so that the sample table and other cached
                    # UI pieces that depend on the reviews map reflect the
                    # newly-saved review immediately in the same session.
                    try:
                        st.experimental_rerun()
                    except Exception:
                        # If rerun isn't available (older Streamlit), ignore.
                        pass

    st.markdown('---')
    st.subheader('Detalhes do pedido')
    st.write('Abra o detalhe de um pedido para checar os campos que vieram no consolidado e comparar com os itens.')
    # allow pre-filling the detail id via ?detail_id=ORDERID so the open-detail
    # action from the table can navigate here and show the detail immediately.
    detail_id = st.text_input('Abrir detalhe por Order ID (cole aqui)', value=str(detail_prefill) if detail_prefill else '')
    if detail_id:
        con = sqlite3.connect(DB_PATH)
        try:
            od = pd.read_sql('SELECT * FROM orders WHERE order_id = ?', con, params=(detail_id,))
            oi = pd.read_sql('SELECT * FROM order_items WHERE order_id = ?', con, params=(detail_id,))
        except Exception:
            od = pd.DataFrame()
            oi = pd.DataFrame()
        con.close()
        st.markdown('**Resumo (visão ML-like)**')
        if not od.empty:
            o = od.iloc[0]
            # Build a breakdown similar to the Mercado Livre detail panel
            breakdown = {
                'Preço do produto': o.get('receita_produtos_brl', 0.0),
                'Tarifa de venda total': o.get('tarifa_venda_impostos_brl', 0.0),
                'Tarifas de envio': o.get('tarifas_envio_brl', 0.0),
                'Cancelamentos e reembolsos': o.get('cancelamentos_reembolsos_brl', 0.0),
                'Dinheiro liberado': o.get('dinheiro_liberado', 0.0),
                'Total (nosso calculado)': o.get('total_brl', 0.0)
            }
            # pretty print the breakdown
            for k, v in breakdown.items():
                if v is None:
                    s = ''
                else:
                    s = fmt_brl_signed(v)
                st.write(f"{k}: ", s)
            # show internal pending heuristics inside a collapsed section with
            # an explicit explanation so users don't mistake it for the ML "Total"
            with st.expander('Valores internos (debug): _valor_pendente / _valor_passivel_extorno'):
                st.write('Estes valores são heurísticos internos sobre possíveis estornos/pendências. Não representam o "Total/Prejuízo" mostrado no painel do Mercado Livre.')
                st.write('_valor_pendente:', fmt_brl(o.get('_valor_pendente', 0.0)))
                st.write('_valor_passivel_extorno:', fmt_brl(o.get('_valor_passivel_extorno', 0.0)))
        else:
            st.info('Pedido não encontrado na tabela `orders`.')
            st.markdown('**Linha consolidada (orders)**')
            if not od.empty:
                st.json(od.to_dict(orient='records')[0])
            else:
                st.info('Pedido não encontrado na tabela `orders`.')

            st.markdown('**Itens (order_items)**')
            if not oi.empty:
                # render items as HTML table to avoid pyarrow
                oi_display = oi.copy()
                if 'preco_unitario' in oi_display.columns:
                    oi_display['preco_unitario'] = oi_display['preco_unitario'].apply(lambda x: fmt_brl(x) if pd.notna(x) else '')
                st.markdown(render_interactive_table(oi_display, table_id='items_tbl'), unsafe_allow_html=True)
            else:
                st.info('Sem itens encontrados para este pedido.')

            # --- Diagnostic: show raw review and action rows for this Order ID ---
            try:
                con2 = sqlite3.connect(DB_PATH)
                try:
                    review_raw = pd.read_sql('SELECT * FROM reviews WHERE order_id = ?', con2, params=(detail_id,))
                except Exception:
                    review_raw = pd.DataFrame()
                try:
                    actions_raw = pd.read_sql('SELECT * FROM actions WHERE order_id = ? ORDER BY id DESC LIMIT 20', con2, params=(detail_id,))
                except Exception:
                    actions_raw = pd.DataFrame()
            finally:
                try:
                    con2.close()
                except Exception:
                    pass

            st.markdown('**Diagnóstico (raw) — reviews / actions para este Order ID**')
            if not review_raw.empty:
                # converter reviewed_at para fuso local antes de mostrar
                review_raw = _convert_ts_for_display(review_raw, ts_cols='reviewed_at')
                st.write('Row raw em `reviews` (colunas: order_id, reviewed, reviewed_by, reviewed_at, review_description)')
                st.dataframe(review_raw)
                # Optional debug: show what the in-memory reviews map contains
                # for this order id. Enable by setting environment variable
                # SHOW_REVIEW_DEBUG=1 in the deployment (safe, opt-in).
                try:
                    if os.environ.get('SHOW_REVIEW_DEBUG', '') == '1':
                        rm = get_reviews_map()
                        st.write('DEBUG reviews_map entry for this order_id:', rm.get(str(detail_id)))
                except Exception:
                    pass
            else:
                st.info('Nenhuma linha encontrada em `reviews` para este Order ID.')

            if not actions_raw.empty:
                # converter created_at para fuso local antes da exibição
                actions_raw = _convert_ts_for_display(actions_raw, ts_cols='created_at')
                st.write('Últimas ações registradas (tabela `actions`)')
                st.dataframe(actions_raw)
            else:
                st.info('Nenhuma ação registrada para este Order ID.')

        st.markdown('---')
        st.subheader('Histórico de ações (últimas 50)')
        con = sqlite3.connect(DB_PATH)
        try:
            actions = pd.read_sql('SELECT id, order_id, user, action, note, created_at FROM actions ORDER BY id DESC LIMIT 50', con)
        except Exception:
            actions = pd.DataFrame()
        con.close()
        if not actions.empty:
            actions_display = actions.copy()
            # converter created_at para fuso local (São Paulo) antes da exibição
            actions_display = _convert_ts_for_display(actions_display, ts_cols='created_at')
            st.markdown(render_interactive_table(actions_display, table_id='actions_tbl'), unsafe_allow_html=True)
        else:
            st.info('Nenhuma ação registrada ainda.')

        # Bulk review marking
        st.markdown('---')
        st.subheader('Marcar Revisado em Lote')
        st.write('Informe múltiplos Order IDs (separados por vírgula ou nova linha) e marque/desmarque como Revisado.')
        with st.form('bulk_review_form'):
            bulk_orders = st.text_area('Order IDs (ex: 2000..., 2001..., ou um por linha)')
            bulk_user = st.text_input('Usuário (quem revisa)', value='operator')
            bulk_reviewed = st.selectbox('Ação', ['Marcar como Revisado', 'Desmarcar Revisado'])
            bulk_submit = st.form_submit_button('Aplicar em lote')
            if bulk_submit:
                ids = [s.strip() for s in bulk_orders.replace(',', '\n').split('\n') if s.strip()]
                if not ids:
                    st.error('Informe pelo menos um Order ID')
                else:
                    for oid in ids:
                        set_review(oid, True if bulk_reviewed=='Marcar como Revisado' else False, bulk_user)
                    st.success(f'{bulk_reviewed} aplicado a {len(ids)} pedidos')

        st.markdown('---')
        st.subheader('Relatórios e exportação')
        col_exp1, col_exp2 = st.columns(2)
        with col_exp1:
            if st.button('Exportar tabela atual para CSV'):
                out_dir = Path('reports')
                out_dir.mkdir(parents=True, exist_ok=True)
                suffix = '_prejuizo' if only_loss else '_full'
                out = out_dir / f'export{suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                # include review columns
                export_df = df.copy()
                reviews_map = get_reviews_map()
                export_df['Revisado'] = export_df['order_id'].apply(lambda oid: bool(reviews_map.get(str(oid), {}).get('reviewed', 0)))
                export_df['Revisado_por'] = export_df['order_id'].apply(lambda oid: reviews_map.get(str(oid), {}).get('reviewed_by'))
                export_df['Revisado_em'] = export_df['order_id'].apply(lambda oid: reviews_map.get(str(oid), {}).get('reviewed_at'))
                # robust formatting: prefer UTC-aware parsing then fallback to naive localization
                try:
                    # Centralized conversion (handles aware/naive values and runtime fallbacks)
                    export_df = _convert_ts_for_display(export_df, ts_cols='Revisado_em')
                    export_df['Revisado_em'] = export_df['Revisado_em'].fillna('')
                except Exception:
                    pass
                # add detail URL for each order so CSV consumers can open the sale detail directly
                export_df['detail_url'] = export_df['order_id'].astype(str).apply(lambda oid: f'https://www.mercadolivre.com.br/vendas/{oid}/detalhe' if oid else '')
                export_df.to_csv(out, index=False, encoding='utf-8-sig')
                st.success(f'Export salvo em {out}')
        with col_exp2:
            if st.button('Exportar tabela atual para XLSX (formatado)'):
                out_dir = Path('reports')
                out_dir.mkdir(parents=True, exist_ok=True)
                suffix = '_prejuizo' if only_loss else '_full'
                out_x = out_dir / f'export{suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                # build friendly display names (Português)
                display_map = {
                    'order_id':'Order ID',
                    'data_venda':'Data da venda',
                    'total_brl':'Total (R$)',
                    '_valor_passivel_extorno':'Passível de estorno (R$)',
                    'dinheiro_liberado':'Dinheiro liberado (R$)',
                    'sku':'SKU',
                    'preco_unitario':'Preço unitário (R$)',
                    'unidades':'Unidades',
                    'resultado':'Resultado'
                }
                export_df = df.copy()
                reviews_map = get_reviews_map()
                export_df['Revisado'] = export_df['order_id'].apply(lambda oid: bool(reviews_map.get(str(oid), {}).get('reviewed', 0)))
                export_df['Revisado_por'] = export_df['order_id'].apply(lambda oid: reviews_map.get(str(oid), {}).get('reviewed_by'))
                export_df['Revisado_em'] = export_df['order_id'].apply(lambda oid: reviews_map.get(str(oid), {}).get('reviewed_at'))
                try:
                    export_df = _convert_ts_for_display(export_df, ts_cols='Revisado_em')
                    export_df['Revisado_em'] = export_df['Revisado_em'].fillna('')
                except Exception:
                    pass
                # rename columns to Portuguese friendly names where possible
                display_names = {c: display_map.get(c, c) for c in export_df.columns}
                export_df.rename(columns=display_names, inplace=True)
                ok, err = create_xlsx_export(export_df, out_x, display_names)
                if ok:
                    st.success(f'Export XLSX salvo em {out_x}')
                else:
                    st.error('Erro ao salvar XLSX: ' + (err or ''))

# Novo: converte colunas de timestamp (ISO/UTC) para America/Sao_Paulo para exibição
def _convert_ts_for_display(df: pd.DataFrame, ts_cols):
    """
    Convert timestamp columns for display (to America/Sao_Paulo).

    Behavior:
      - Prefer to parse all values as UTC where possible (this matches how
        `set_review` stores timestamps: UTC-aware ISO strings).
      - For values that can't be parsed as UTC, fall back to parsing as naive
        datetimes and localizing them according to the environment setting
        `NAIVE_TIMESTAMP_INTERPRETATION` which may be 'UTC' or 'LOCAL'.

    Rationale: production contains a mix of naive and aware timestamps. Using
    a single configurable rule avoids surprises; default is to treat naive as
    UTC so stored UTC-like naive strings display correctly in America/Sao_Paulo.
    """
    if df is None or df.empty:
        return df
    if isinstance(ts_cols, str):
        ts_cols = [ts_cols]

    # how to interpret naive timestamps: 'UTC' or 'LOCAL' (America/Sao_Paulo)
    naive_mode = os.environ.get('NAIVE_TIMESTAMP_INTERPRETATION', 'UTC').upper()
    for c in ts_cols:
        if c not in df.columns:
            continue

        try:
            s = df[c]
            # First, try parsing everything as UTC (this will interpret naive
            # strings as UTC which is the safer default when timestamps were
            # produced by servers/helpers that already used UTC).
            parsed = pd.to_datetime(s, utc=True, errors='coerce')

            # For any values that failed parsing as UTC, try a naive parse and
            # then localize according to naive_mode.
            not_parsed = parsed.isna()
            if not_parsed.any():
                fallback = pd.to_datetime(s[not_parsed], errors='coerce')
                if not fallback.empty:
                    if naive_mode == 'UTC':
                        try:
                            fallback = fallback.dt.tz_localize('UTC')
                        except Exception:
                            # If localize fails, leave as NaT
                            fallback = pd.Series([pd.NaT] * len(fallback), index=fallback.index)
                    else:
                        try:
                            fallback = fallback.dt.tz_localize('America/Sao_Paulo')
                        except Exception:
                            fallback = pd.Series([pd.NaT] * len(fallback), index=fallback.index)
                    # assign back
                    parsed.loc[not_parsed] = fallback

            # Finally, convert all tz-aware times to America/Sao_Paulo and
            # format for display. Any remaining NaT values become empty strings.
            try:
                # Preferred: convert tz-aware series to America/Sao_Paulo
                converted = parsed.dt.tz_convert('America/Sao_Paulo')
                df[c] = converted.dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
            except Exception:
                # If tz conversion fails (for example tzdata not available in the
                # runtime), fall back to a safe manual shift: interpret values as
                # UTC then subtract 3 hours to approximate America/Sao_Paulo.
                # Some minimal Python runtimes (e.g. musl builds or stripped
                # containers) may not include full tzdata and tz_convert calls
                # can raise. In that case, avoid relying on tz-aware ops and
                # perform a best-effort manual shift: parse as naive datetimes
                # (interpreting them as UTC) and subtract 3 hours.
                    try:
                        # Manual, tzdata-free fallback using dateutil.parser which
                        # handles both aware (with offset) and naive strings. This
                        # avoids relying on pandas tz_convert which can fail in
                        # minimal runtimes without tzdb. We interpret naive
                        # timestamps according to `naive_mode` (UTC or LOCAL).
                        from dateutil import parser as _parser
                        from datetime import timedelta

                        def _fmt_manual(x):
                            try:
                                if x is None:
                                    return ''
                                s_val = str(x)
                                if not s_val or s_val.lower() in ('nan', 'none'):
                                    return ''
                                dt = _parser.parse(s_val)
                                # If parsed value has no tzinfo, interpret per naive_mode
                                if getattr(dt, 'tzinfo', None) is None:
                                    if naive_mode == 'UTC':
                                        # treat naive as UTC, convert to Sao_Paulo by -3h
                                        dt_utc = dt.replace(tzinfo=timezone.utc)
                                        dt_sp = dt_utc - timedelta(hours=3)
                                    else:
                                        # treat naive as already local (America/Sao_Paulo)
                                        dt_sp = dt
                                else:
                                    # aware datetime: normalize to UTC then shift -3h
                                    dt_utc = dt.astimezone(timezone.utc)
                                    dt_sp = dt_utc - timedelta(hours=3)
                                return dt_sp.strftime('%Y-%m-%d %H:%M:%S')
                            except Exception:
                                return ''

                        df[c] = s.map(_fmt_manual)
                    except Exception:
                        # Last resort: format whatever could be parsed as naive
                        df[c] = pd.to_datetime(df[c], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
        except Exception:
            # Last resort: fallback to naive parse and string formatting
            try:
                df[c] = pd.to_datetime(df[c], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
            except Exception:
                df[c] = df[c].astype(str).fillna('')

    return df

if __name__ == '__main__':
    main()
